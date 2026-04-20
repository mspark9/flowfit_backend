"""
HR 인사 평가 분석 서비스 — DB 기반 AI 리포트

설계 원칙 (sales_performance_service.py 동일):
- 수치·이상감지·등급은 Python 규칙으로 '결정적'으로 계산
- LLM은 요약·원인 추정·액션 추천 등 '해석' 역할만 담당
"""
import json
from io import BytesIO

from openai import OpenAI
from config import settings
from services.HR.hr_evaluation_entry_service import (
    fetch_evaluation,
    DEFAULT_CRITERIA,
    EVAL_SLOTS,
    _safe_float,
    _safe_int,
)


def _enabled_items(criteria: dict) -> list:
    return [i for i in (criteria or {}).get("items", []) if i.get("enabled")]


def _combined_score(ind: dict, criteria: dict) -> float:
    enabled = _enabled_items(criteria)
    total_w = sum(_safe_float(i.get("weight", 0)) for i in enabled) or 1.0
    s = 0.0
    for it in enabled:
        raw = _safe_float(ind.get(it["key"], 0))
        cap = _safe_float(it.get("max", 100)) or 100.0
        norm = min(100.0, raw / cap * 100.0)
        s += norm * (_safe_float(it.get("weight", 0)) / total_w)
    return round(s, 1)

client = OpenAI(api_key=settings.openai_api_key)

# ────────────────────────────────────────────────────────────
# 이상감지 룰 임계값
# ────────────────────────────────────────────────────────────

ANOMALY_RULES = {
    "budget_overrun_pct":      110.0,   # 예산 집행률 > 110% → 초과 경고
    "budget_underuse_pct":      50.0,   # 예산 집행률 < 50% → 미활용 주의
    "low_target_achievement":   60.0,   # 목표 달성률 < 60% → 부서 저성과
    "low_project_completion":   60.0,   # 프로젝트 완수율 < 60% → 주의
    "low_individual_score":     50.0,   # 개인 종합 점수 < 50 → D등급 주의
    "competency_gap":           25.0,   # 업무-역량 점수 차이 > 25 → 불균형
    "high_d_ratio_pct":         20.0,   # D등급 비율 > 20% → 부서 주의
}


def _detect_anomalies(departments: list, individuals: list, criteria: dict) -> list:
    anomalies = []

    # 부서별 이상감지
    for d in departments:
        dept_name = d["department"]
        exec_rate = d.get("budget_execution_rate", 0)
        target = d.get("target_achievement", 0)
        proj = d.get("project_completion", 0)

        if exec_rate > ANOMALY_RULES["budget_overrun_pct"]:
            anomalies.append({
                "type": "주의", "item": f"{dept_name} 예산 초과",
                "detail": f"예산 집행률 {exec_rate:.1f}% (임계 {ANOMALY_RULES['budget_overrun_pct']}% 초과)",
                "severity": "높음" if exec_rate > 130 else "중간",
            })
        elif 0 < exec_rate < ANOMALY_RULES["budget_underuse_pct"]:
            anomalies.append({
                "type": "주의", "item": f"{dept_name} 예산 미활용",
                "detail": f"예산 집행률 {exec_rate:.1f}% (임계 {ANOMALY_RULES['budget_underuse_pct']}% 미만)",
                "severity": "중간",
            })

        if 0 < target < ANOMALY_RULES["low_target_achievement"]:
            anomalies.append({
                "type": "급락", "item": f"{dept_name} 목표 달성률 저조",
                "detail": f"목표 달성률 {target:.1f}% (임계 {ANOMALY_RULES['low_target_achievement']}% 미만)",
                "severity": "높음" if target < 40 else "중간",
            })

        if 0 < proj < ANOMALY_RULES["low_project_completion"]:
            anomalies.append({
                "type": "주의", "item": f"{dept_name} 프로젝트 완수율 저조",
                "detail": f"프로젝트 완수율 {proj:.1f}%",
                "severity": "중간",
            })

        # 부서 내 D등급 비율
        dept_inds = [i for i in individuals if i["department"] == dept_name]
        if dept_inds:
            d_count = sum(1 for i in dept_inds if i.get("overall_grade") == "D")
            d_ratio = d_count / len(dept_inds) * 100
            if d_ratio > ANOMALY_RULES["high_d_ratio_pct"]:
                anomalies.append({
                    "type": "주의", "item": f"{dept_name} 저성과 인원 비율 높음",
                    "detail": f"D등급 {d_count}명 / 전체 {len(dept_inds)}명 ({d_ratio:.0f}%)",
                    "severity": "중간",
                })

    # 개인별 이상감지 — criteria 기반 종합 점수
    for ind in individuals:
        name = ind["employee_name"]
        combined = _combined_score(ind, criteria)

        if combined < ANOMALY_RULES["low_individual_score"]:
            anomalies.append({
                "type": "주의", "item": f"{name} 종합 점수 저조",
                "detail": f"종합 {combined:.1f}점",
                "severity": "높음" if combined < 35 else "중간",
            })

    return anomalies


# ────────────────────────────────────────────────────────────
# LLM 프롬프트
# ────────────────────────────────────────────────────────────

EVALUATION_PROMPT = """
당신은 기업 인사 평가 분석 전문가입니다.
아래 평가 데이터와 '규칙 기반으로 사전 감지된 이상 항목'을 바탕으로
인사팀장 보고용 종합 평가 리포트를 생성하세요.

[중요] 이상 항목(anomalies)은 이미 결정적으로 계산되었습니다.
당신의 역할은 각 이상 항목에 대해 'cause'(추정 원인, 1~2문장)만 추가하는 것입니다.
type/item/detail/severity는 원본 그대로 유지하세요.

평가 기간: {eval_label}
평가 대상: {department_scope}

부서별 KPI:
{dept_text}

개인 평가 (종합점수 상위순):
{individual_text}

등급 분포:
{grade_dist_text}

사전 감지된 이상 항목:
{anomalies_text}

JSON으로만 응답하세요:
{{
  "summary": "평가 결과 3줄 요약 (인사팀장 보고용, 핵심 수치 포함)",
  "department_insights": [
    {{"department": "부서명", "insight": "부서별 평가 코멘트 (2~3문장)"}}
  ],
  "anomaly_causes": [
    {{"item": "이상 항목의 item과 동일", "cause": "원인 추정 (1~2문장)"}}
  ],
  "top_performers": "우수 성과자 분석 (2~3문장, 구체적 이름과 점수 언급)",
  "improvement_areas": "개선 필요 영역 분석 (2~3문장)",
  "recommendations": ["액션 추천 1", "액션 추천 2", "액션 추천 3"]
}}
"""


def analyze_evaluation(eval_key: str, department: str = "") -> dict:
    data = fetch_evaluation(eval_key)
    if not data:
        raise ValueError(f"등록된 평가가 없습니다: {eval_key}")

    period = data["period"]
    criteria = period.get("criteria_config") or {}
    enabled = _enabled_items(criteria)
    all_depts = data["departments"]
    all_individuals = data["individuals"]

    # 부서 필터
    if department:
        depts = [d for d in all_depts if d["department"] == department]
        individuals = [i for i in all_individuals if i["department"] == department]
    else:
        depts = all_depts
        individuals = all_individuals

    if not individuals:
        raise ValueError("평가 대상 직원이 없습니다.")

    # 등급 분포 계산
    grade_dist = {"A": 0, "B": 0, "C": 0, "D": 0}
    for ind in individuals:
        g = ind.get("overall_grade", "D")
        if g in grade_dist:
            grade_dist[g] += 1

    # 이상감지
    anomalies = _detect_anomalies(depts, individuals, criteria)

    # LLM 텍스트 구성
    dept_text = "\n".join(
        f"- {d['department']}: 목표달성률 {d['target_achievement']:.1f}% / "
        f"프로젝트완수율 {d['project_completion']:.1f}% / "
        f"협업점수 {d['collaboration_score']:.1f} / "
        f"예산집행률 {d['budget_execution_rate']:.1f}% / "
        f"인원 {d['headcount']}명"
        for d in depts
    ) or "- (등록된 부서 없음)"

    sorted_inds = sorted(
        individuals,
        key=lambda x: _combined_score(x, criteria),
        reverse=True,
    )

    def _ind_line(i: dict) -> str:
        scores_part = " ".join(
            f"{it['label']}:{_safe_float(i.get(it['key'], 0)):.1f}" for it in enabled
        )
        return (
            f"- {i['employee_name']}({i['department']}/{i.get('position','')}) "
            f"{scores_part} 종합:{_combined_score(i, criteria):.1f} "
            f"등급:{i.get('overall_grade','')}"
        )

    individual_text = "\n".join(_ind_line(i) for i in sorted_inds) or "- (등록된 개인 평가 없음)"

    grade_dist_text = " / ".join(f"{g}등급: {c}명" for g, c in grade_dist.items())

    anomalies_text = "\n".join(
        f"- [{a['type']}] {a['item']} — {a['detail']} (심각도: {a['severity']})"
        for a in anomalies
    ) or "- (사전 감지된 이상 항목 없음)"

    department_scope = department if department else "전체 부서"

    res = client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[{
            "role": "user",
            "content": EVALUATION_PROMPT.format(
                eval_label=period["eval_label"],
                department_scope=department_scope,
                dept_text=dept_text,
                individual_text=individual_text,
                grade_dist_text=grade_dist_text,
                anomalies_text=anomalies_text,
            ),
        }],
        max_tokens=1500,
    )

    llm_result = json.loads(res.choices[0].message.content)

    # LLM의 cause를 결정적 anomalies에 병합
    cause_map = {
        c.get("item", ""): c.get("cause", "")
        for c in (llm_result.get("anomaly_causes") or [])
    }
    for a in anomalies:
        a["cause"] = cause_map.get(a["item"], "")

    # 통계 — 활성 항목별 평균 + 종합 평균
    total = len(individuals)
    item_averages = []
    for it in enabled:
        avg = sum(_safe_float(i.get(it["key"], 0)) for i in individuals) / total if total else 0
        item_averages.append({"key": it["key"], "label": it["label"], "average": round(avg, 1)})
    avg_combined = sum(_combined_score(i, criteria) for i in individuals) / total if total else 0

    return {
        "period": period,
        "criteria": criteria,
        "metrics": {
            "total_employees": total,
            "avg_combined_score": round(avg_combined, 1),
            "item_averages": item_averages,
            "grade_distribution": grade_dist,
        },
        "departments": depts,
        "individuals": [
            {
                "employee_name": i["employee_name"],
                "department": i["department"],
                "position": i.get("position", ""),
                **{slot: i.get(slot, 0) for slot in (it["key"] for it in enabled)},
                "combined_score": _combined_score(i, criteria),
                "overall_grade": i.get("overall_grade", ""),
            }
            for i in sorted_inds
        ],
        "anomalies": anomalies,
        "summary": llm_result.get("summary", ""),
        "department_insights": llm_result.get("department_insights", []),
        "top_performers": llm_result.get("top_performers", ""),
        "improvement_areas": llm_result.get("improvement_areas", ""),
        "recommendations": llm_result.get("recommendations", []),
    }


# ────────────────────────────────────────────────────────────
# Excel 내보내기
# ────────────────────────────────────────────────────────────

def export_evaluation_to_excel(eval_key: str, department: str = "") -> bytes:
    report = analyze_evaluation(eval_key, department=department)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3B82F6")  # blue-500
    center = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, headers):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    # 1) 요약 시트
    ws_sum = wb.active
    ws_sum.title = "요약"
    _write_header(ws_sum, ["항목", "값"])
    p = report["period"]
    m = report["metrics"]
    ws_sum.append(["평가 기간", p["eval_label"]])
    ws_sum.append(["평가 유형", p["eval_type"]])
    ws_sum.append(["시작일", p["start_date"]])
    ws_sum.append(["종료일", p["end_date"]])
    ws_sum.append(["평가 대상 인원", m["total_employees"]])
    ws_sum.append(["평균 종합 점수", m.get("avg_combined_score", 0)])
    for ia in m.get("item_averages", []):
        ws_sum.append([f"평균 {ia['label']}", ia["average"]])
    ws_sum.append([])
    ws_sum.append(["등급 분포"])
    for grade, count in m["grade_distribution"].items():
        ws_sum.append([f"  {grade}등급", f"{count}명"])
    ws_sum.append([])
    ws_sum.append(["종합 요약", report.get("summary", "")])
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 60

    # 2) 부서 KPI 시트
    ws_dept = wb.create_sheet("부서 KPI")
    _write_header(ws_dept, [
        "부서", "목표달성률(%)", "프로젝트완수율(%)", "협업점수",
        "예산집행률(%)", "예산총액", "예산집행액", "인원",
    ])
    for d in report["departments"]:
        ws_dept.append([
            d["department"],
            d.get("target_achievement", 0),
            d.get("project_completion", 0),
            d.get("collaboration_score", 0),
            d.get("budget_execution_rate", 0),
            d.get("budget_total", 0),
            d.get("budget_spent", 0),
            d.get("headcount", 0),
        ])
    ws_dept.column_dimensions["A"].width = 20
    for col in "BCDEFGH":
        ws_dept.column_dimensions[col].width = 16

    # 3) 개인 평가 시트 — 활성 평가 항목별 동적 컬럼
    ws_ind = wb.create_sheet("개인 평가")
    enabled_items = _enabled_items(report.get("criteria") or {})
    headers = ["이름", "부서", "직급"] + [it["label"] for it in enabled_items] + ["종합점수", "종합등급"]
    _write_header(ws_ind, headers)
    for i in report["individuals"]:
        row = [i["employee_name"], i["department"], i.get("position", "")]
        for it in enabled_items:
            row.append(i.get(it["key"], 0))
        row.append(i.get("combined_score", 0))
        row.append(i.get("overall_grade", ""))
        ws_ind.append(row)
    ws_ind.column_dimensions["A"].width = 14
    ws_ind.column_dimensions["B"].width = 18

    # 4) 이상감지 시트
    ws_anom = wb.create_sheet("이상감지")
    _write_header(ws_anom, ["유형", "항목", "세부", "심각도", "원인 추정"])
    for a in report.get("anomalies", []):
        ws_anom.append([
            a.get("type", ""), a.get("item", ""), a.get("detail", ""),
            a.get("severity", ""), a.get("cause", ""),
        ])
    for col, width in zip("ABCDE", [10, 28, 40, 10, 50]):
        ws_anom.column_dimensions[col].width = width

    # 5) AI 인사이트 시트
    ws_ai = wb.create_sheet("AI인사이트")
    _write_header(ws_ai, ["항목", "내용"])
    ws_ai.append(["우수 성과자", report.get("top_performers", "")])
    ws_ai.append(["개선 필요 영역", report.get("improvement_areas", "")])
    ws_ai.append([])
    ws_ai.append(["부서별 인사이트"])
    for di in report.get("department_insights", []):
        ws_ai.append([di.get("department", ""), di.get("insight", "")])
    ws_ai.append([])
    ws_ai.append(["액션 추천"])
    for idx, rec in enumerate(report.get("recommendations", []), start=1):
        ws_ai.append([f"{idx}", rec])
    ws_ai.column_dimensions["A"].width = 16
    ws_ai.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
