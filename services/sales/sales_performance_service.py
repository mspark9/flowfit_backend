"""
영업 실적 분석 서비스 — DB(PostgreSQL) 기반 AI 리포트

설계 원칙:
- 수치·이상감지·전환율은 Python 규칙으로 '결정적'으로 계산
- LLM은 요약·원인 추정·액션 추천 등 '해석' 역할만 담당
- 데이터 원본은 sales_period_summary / sales_pipeline_stages / sales_member_performance
"""
import json
from io import BytesIO

from openai import OpenAI
from config import settings
from services.sales.sales_performance_entry_service import (
    fetch_performance,
    list_periods,
    previous_period_key,
)

client = OpenAI(api_key=settings.openai_api_key)

# 이상감지 룰 임계값 (팀 합의 시 config화 가능)
ANOMALY_RULES = {
    "target_shortfall_pct":     -20.0,  # 목표 대비 N% 이상 미달 → 급락
    "target_overshoot_pct":     +20.0,  # 목표 대비 N% 이상 초과 → 급등
    "growth_drop_pct":          -15.0,  # 전기 대비 N% 이상 하락 → 급락
    "growth_surge_pct":         +25.0,  # 전기 대비 N% 이상 상승 → 급등
    "win_rate_low_pct":          30.0,  # 수주율 N% 미만 → 주의
    "conversion_bottleneck_pct": 25.0,  # 단계 전환율 N% 미만 → 주의 (병목)
    "member_zero_win_deals":        2,  # 수주 0 & 진행 딜 N건 이상 → 주의
}


def _calc_conversion_rates(pipeline: list) -> list:
    """
    인접 파이프라인 단계 간 전환율을 계산합니다.

    Args:
        pipeline: [{stage_name/ stage, stage_count/ count, ...}, ...]
                  DB 조회 키(stage_name, stage_count)와 API 응답 키(stage, count) 모두 허용.
    Returns:
        [ { from, to, rate: float(%) }, ... ]
    """
    def _name(s): return s.get("stage_name") or s.get("stage") or ""
    def _count(s): return s.get("stage_count") if "stage_count" in s else s.get("count", 0)

    rates = []
    for i in range(len(pipeline) - 1):
        prev_count = _count(pipeline[i]) or 0
        next_count = _count(pipeline[i + 1]) or 0
        rate = (next_count / prev_count * 100) if prev_count > 0 else 0.0
        rates.append({
            "from": _name(pipeline[i]),
            "to":   _name(pipeline[i + 1]),
            "rate": round(rate, 1),
        })
    return rates


def _detect_anomalies(
    achievement_rate: float,
    growth_rate: float,
    win_rate: float,
    conversion_rates: list,
    filtered_members: list,
) -> list:
    """
    규칙 기반 이상감지 — LLM 호출 전에 결정적으로 수행.

    filtered_members: [{member_name, revenue, deals, wins}, ...]
    """
    anomalies: list = []

    # 1. 목표 대비
    diff_target = achievement_rate - 100
    if diff_target <= ANOMALY_RULES["target_shortfall_pct"]:
        anomalies.append({
            "type":     "급락",
            "item":     "목표 매출 미달",
            "detail":   f"목표 달성률 {achievement_rate:.1f}% — 목표 대비 {diff_target:+.1f}%p",
            "severity": "높음" if diff_target <= -30 else "중간",
        })
    elif diff_target >= ANOMALY_RULES["target_overshoot_pct"]:
        anomalies.append({
            "type":     "급등",
            "item":     "목표 매출 초과 달성",
            "detail":   f"목표 달성률 {achievement_rate:.1f}% — 목표 대비 {diff_target:+.1f}%p",
            "severity": "중간",
        })

    # 2. 전기 대비
    if growth_rate <= ANOMALY_RULES["growth_drop_pct"]:
        anomalies.append({
            "type":     "급락",
            "item":     "전기 대비 매출 하락",
            "detail":   f"전기 대비 {growth_rate:+.1f}%",
            "severity": "높음" if growth_rate <= -25 else "중간",
        })
    elif growth_rate >= ANOMALY_RULES["growth_surge_pct"]:
        anomalies.append({
            "type":     "급등",
            "item":     "전기 대비 매출 급등",
            "detail":   f"전기 대비 +{growth_rate:.1f}%",
            "severity": "중간",
        })

    # 3. 수주율
    if win_rate < ANOMALY_RULES["win_rate_low_pct"]:
        anomalies.append({
            "type":     "주의",
            "item":     "낮은 수주율",
            "detail":   f"수주율 {win_rate:.1f}% (임계 {ANOMALY_RULES['win_rate_low_pct']}% 미만)",
            "severity": "중간",
        })

    # 4. 파이프라인 전환율 병목
    for cr in conversion_rates:
        if cr["rate"] < ANOMALY_RULES["conversion_bottleneck_pct"]:
            anomalies.append({
                "type":     "주의",
                "item":     f"전환 병목: {cr['from']} → {cr['to']}",
                "detail":   f"전환율 {cr['rate']}% (임계 {ANOMALY_RULES['conversion_bottleneck_pct']}% 미만)",
                "severity": "중간",
            })

    # 5. 팀원별 — 진행 딜이 있는데 수주 0
    for m in filtered_members:
        name = m.get("member_name") or m.get("name") or ""
        wins  = int(m.get("wins", 0))
        deals = int(m.get("deals", 0))
        if wins == 0 and deals >= ANOMALY_RULES["member_zero_win_deals"]:
            anomalies.append({
                "type":     "주의",
                "item":     f"{name} 수주 실적 부진",
                "detail":   f"진행 {deals}건 중 수주 0건",
                "severity": "중간",
            })

    return anomalies


PERFORMANCE_PROMPT = """
당신은 영업 데이터 분석 전문가입니다.
아래 CRM 실적 데이터와 '규칙 기반으로 사전 감지된 이상 항목'을 바탕으로
영업팀장 보고용 리포트를 생성하세요.

[중요] 이상 항목(anomalies)은 이미 결정적으로 계산되었습니다.
당신의 역할은 각 이상 항목에 대해 'cause'(추정 원인, 1~2문장)만 추가하는 것입니다.
type/item/detail/severity는 원본 그대로 유지하세요.

분석 기간: {period_label}
목표 매출: {target_revenue:,}원
실제 매출: {actual_revenue:,}원
목표 달성률: {achievement_rate:.1f}%
전기 대비: {growth_rate:+.1f}%
수주율: {win_rate:.1f}%

파이프라인 현황 (단계별 건수·금액):
{pipeline_text}

파이프라인 전환율 (인접 단계):
{conversion_text}

팀원별 실적:
{member_text}

사전 감지된 이상 항목:
{anomalies_text}

JSON으로만 응답하세요:
{{
  "summary": "실적 3줄 요약 (팀장 보고용, 핵심 수치 포함)",
  "achievement_comment": "목표 달성률에 대한 평가 코멘트 (1~2문장)",
  "anomaly_causes": [
    {{"item": "이상 항목의 item과 동일", "cause": "원인 추정 (1~2문장)"}}
  ],
  "pipeline_insight": "파이프라인 분석 코멘트 — 전환율과 병목을 구체적으로 언급 (3~4문장)",
  "top_performer": "최고 실적자 및 코멘트 (1~2문장)",
  "risk_deals": "리스크 딜 분석 (협상 중 딜의 전환 가능성, 2~3문장)",
  "recommendations": ["액션 추천 1", "액션 추천 2", "액션 추천 3"]
}}
"""


def get_team_members(period_key: str = "") -> list:
    """
    팀원 목록을 반환합니다.
    - period_key 가 주어지면 해당 기간에 등록된 팀원만 반환 (동적).
    - 비어 있으면 모든 기간을 합쳐 고유 팀원 이름 목록 반환.
    첫 번째 항목은 '팀 전체'(id='all').
    """
    from database import get_connection
    conn = get_connection()
    try:
        cur = conn.cursor()
        if period_key:
            cur.execute(
                "SELECT DISTINCT member_name FROM sales_member_performance WHERE period_key=%s ORDER BY member_name",
                (period_key,),
            )
        else:
            cur.execute("SELECT DISTINCT member_name FROM sales_member_performance ORDER BY member_name")
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    members = [{"id": "all", "name": "팀 전체"}]
    for r in rows:
        name = r[0]
        if name:
            # id는 member_name 자체 사용 (한글 name → id 매핑 단순화)
            members.append({"id": name, "name": name})
    return members


def get_periods(period_type: str = "") -> list:
    """등록된 분석 가능 기간 목록 (최신순). /performance/periods 전용."""
    return list_periods(period_type=period_type)


def get_performance_trend(period_type: str, limit: int = 6) -> list:
    """
    지정 기간 타입의 최근 N개 기간 지표 추세를 반환합니다.
    (분석 페이지의 추세 차트용 — 달성률·성장률 비교)

    Returns: [
        { period_key, period_label, start_date, end_date,
          target_revenue, actual_revenue, prev_revenue,
          achievement_rate, growth_rate, win_rate },
        ...
    ]  # start_date 오름차순 (차트 X축 왼쪽 = 과거)
    """
    from database import get_connection

    if period_type not in {"month", "quarter", "year"}:
        raise ValueError("period_type은 'month', 'quarter', 'year' 중 하나여야 합니다.")
    if limit <= 0:
        limit = 6

    conn = get_connection()
    try:
        cur = conn.cursor()
        cur.execute(
            """
            SELECT period_key, period_label, start_date, end_date,
                   target_revenue, actual_revenue, prev_revenue,
                   deal_count, win_count
            FROM sales_period_summary
            WHERE period_type = %s
            ORDER BY start_date DESC, period_key DESC
            LIMIT %s
            """,
            (period_type, limit),
        )
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    # 최신순 → 차트 표시는 과거→현재이므로 뒤집어 반환
    rows = list(reversed(rows))

    result = []
    for r in rows:
        target = int(r[4] or 0)
        actual = int(r[5] or 0)
        prev   = int(r[6] or 0)
        deal   = int(r[7] or 0)
        win    = int(r[8] or 0)
        achievement = round((actual / target * 100), 1) if target > 0 else 0.0
        growth      = round(((actual - prev) / prev * 100), 1) if prev > 0 else 0.0
        win_rate    = round((win / deal * 100), 1) if deal > 0 else 0.0

        result.append({
            "period_key":       r[0],
            "period_label":     r[1],
            "start_date":       r[2].isoformat() if r[2] else "",
            "end_date":         r[3].isoformat() if r[3] else "",
            "target_revenue":   target,
            "actual_revenue":   actual,
            "prev_revenue":     prev,
            "deal_count":       deal,
            "win_count":        win,
            "achievement_rate": achievement,
            "growth_rate":      growth,
            "win_rate":         win_rate,
        })
    return result


def get_performance_compare(period_key: str) -> dict:
    """
    선택 기간과 직전 기간의 요약 지표를 나란히 반환합니다.
    - 직전 기간이 DB에 없으면 previous=None
    - 기간 타입(월/분기/연)에 따라 자동으로 직전 기간 key 산출

    Returns: {
        current:  { period_key, period_label, target, actual, achievement_rate, growth_rate, win_rate, deal_count, win_count } | None,
        previous: { ...동일... } | None,
        delta:    { actual_pct, achievement_pct_point, win_rate_pct_point } | None
    }
    """
    def _metrics(data: dict | None) -> dict | None:
        if not data:
            return None
        s = data["summary"]
        target = int(s["target_revenue"] or 0)
        actual = int(s["actual_revenue"] or 0)
        prev   = int(s["prev_revenue"] or 0)
        deal   = int(s["deal_count"] or 0)
        win    = int(s["win_count"] or 0)
        achievement = round((actual / target * 100), 1) if target > 0 else 0.0
        growth      = round(((actual - prev) / prev * 100), 1) if prev > 0 else 0.0
        win_rate    = round((win / deal * 100), 1) if deal > 0 else 0.0
        return {
            "period_key":       s["period_key"],
            "period_label":     s["period_label"],
            "target_revenue":   target,
            "actual_revenue":   actual,
            "achievement_rate": achievement,
            "growth_rate":      growth,
            "win_rate":         win_rate,
            "deal_count":       deal,
            "win_count":        win,
        }

    current = _metrics(fetch_performance(period_key))
    if not current:
        raise ValueError(f"등록된 실적이 없습니다: {period_key}")

    prev_key = previous_period_key(period_key)
    previous = _metrics(fetch_performance(prev_key)) if prev_key else None

    delta = None
    if previous:
        cur_actual = current["actual_revenue"]
        prev_actual = previous["actual_revenue"]
        actual_pct = round(((cur_actual - prev_actual) / prev_actual * 100), 1) if prev_actual > 0 else 0.0
        delta = {
            "actual_pct":              actual_pct,
            "achievement_pct_point":   round(current["achievement_rate"] - previous["achievement_rate"], 1),
            "win_rate_pct_point":      round(current["win_rate"] - previous["win_rate"], 1),
        }

    return {
        "current":  current,
        "previous": previous,
        "previous_key": prev_key,
        "delta":    delta,
    }


def analyze_performance(period_key: str, member_id: str) -> dict:
    """
    DB 기반 실적 분석 리포트를 생성합니다.

    Args:
        period_key: sales_period_summary.period_key 값
        member_id:  'all' 또는 member_name 값

    Raises:
        ValueError: period_key 가 존재하지 않을 때
    """
    data = fetch_performance(period_key)
    if not data:
        raise ValueError(f"등록된 실적이 없습니다: {period_key}")

    summary  = data["summary"]
    pipeline = data["pipeline"]
    all_members = data["members"]

    # 결정적 지표 계산
    target = summary["target_revenue"]
    actual = summary["actual_revenue"]
    prev   = summary["prev_revenue"]
    deal_count = summary["deal_count"]
    win_count  = summary["win_count"]

    achievement_rate = (actual / target * 100) if target > 0 else 0.0
    growth_rate      = ((actual - prev) / prev * 100) if prev > 0 else 0.0
    win_rate         = (win_count / deal_count * 100) if deal_count > 0 else 0.0

    conversion_rates = _calc_conversion_rates(pipeline)

    # 팀원 필터링
    if member_id == "all" or not member_id:
        filtered = all_members
    else:
        filtered = [m for m in all_members if m["member_name"] == member_id] or all_members

    # 규칙 기반 이상감지
    anomalies = _detect_anomalies(
        achievement_rate=achievement_rate,
        growth_rate=growth_rate,
        win_rate=win_rate,
        conversion_rates=conversion_rates,
        filtered_members=filtered,
    )

    # LLM 호출용 텍스트
    pipeline_text = "\n".join(
        f"- {p['stage_name']}: {p['stage_count']}건 / {p['stage_amount']:,}원"
        for p in pipeline
    ) or "- (등록된 단계 없음)"
    conversion_text = "\n".join(
        f"- {cr['from']} → {cr['to']}: {cr['rate']}%"
        for cr in conversion_rates
    ) or "- (전환율 없음)"
    member_text = "\n".join(
        f"- {m['member_name']}: 매출 {m['revenue']:,}원 / 수주 {m['wins']}건 / 진행 {m['deals']}건"
        for m in filtered
    ) or "- (등록된 팀원 없음)"
    anomalies_text = "\n".join(
        f"- [{a['type']}] {a['item']} — {a['detail']} (심각도: {a['severity']})"
        for a in anomalies
    ) or "- (사전 감지된 이상 항목 없음)"

    res = client.chat.completions.create(
        model="gpt-4o-mini",
        response_format={"type": "json_object"},
        messages=[{
            "role": "user",
            "content": PERFORMANCE_PROMPT.format(
                period_label=summary["period_label"],
                target_revenue=target,
                actual_revenue=actual,
                achievement_rate=achievement_rate,
                growth_rate=growth_rate,
                win_rate=win_rate,
                pipeline_text=pipeline_text,
                conversion_text=conversion_text,
                member_text=member_text,
                anomalies_text=anomalies_text,
            ),
        }],
        max_tokens=1200,
    )

    llm_result = json.loads(res.choices[0].message.content)

    # LLM의 cause를 결정적 anomalies에 병합
    cause_map = {
        c.get("item", ""): c.get("cause", "")
        for c in (llm_result.get("anomaly_causes") or [])
    }
    for a in anomalies:
        a["cause"] = cause_map.get(a["item"], "")

    # UI 호환 형태로 pipeline/members 변환 (기존 PerformancePage가 stage/count/amount, name을 기대)
    pipeline_ui = [
        {
            "stage":  p["stage_name"],
            "count":  p["stage_count"],
            "amount": p["stage_amount"],
        }
        for p in pipeline
    ]
    members_ui = [
        {
            "name":    m["member_name"],
            "revenue": m["revenue"],
            "deals":   m["deals"],
            "wins":    m["wins"],
        }
        for m in filtered
    ]

    return {
        "metrics": {
            "period":           summary["period_label"],
            "period_key":       summary["period_key"],
            "target_revenue":   target,
            "actual_revenue":   actual,
            "achievement_rate": round(achievement_rate, 1),
            "growth_rate":      round(growth_rate, 1),
            "deal_count":       deal_count,
            "win_count":        win_count,
            "win_rate":         round(win_rate, 1),
        },
        "pipeline":            pipeline_ui,
        "conversion_rates":    conversion_rates,
        "members":             members_ui,
        "anomalies":           anomalies,
        "summary":             llm_result.get("summary", ""),
        "achievement_comment": llm_result.get("achievement_comment", ""),
        "pipeline_insight":    llm_result.get("pipeline_insight", ""),
        "top_performer":       llm_result.get("top_performer", ""),
        "risk_deals":          llm_result.get("risk_deals", ""),
        "recommendations":     llm_result.get("recommendations", []),
    }


def export_performance_to_excel(period_key: str, member_id: str = "all") -> bytes:
    """
    실적 분석 결과를 Excel(xlsx) 바이트로 반환합니다.
    시트 구성: 요약 / 파이프라인 / 팀원 / 이상감지 / AI인사이트
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    report = analyze_performance(period_key=period_key, member_id=member_id)
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="F59E0B")  # amber-500
    center = Alignment(horizontal="center", vertical="center")

    def _write_header(ws, headers):
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

    # 1) 요약 시트
    ws_sum = wb.active
    ws_sum.title = "요약"
    _write_header(ws_sum, ["항목", "값"])
    m = report["metrics"]
    ws_sum.append(["분석 기간",      m["period"]])
    ws_sum.append(["Period Key",     m["period_key"]])
    ws_sum.append(["목표 매출",      m["target_revenue"]])
    ws_sum.append(["실제 매출",      m["actual_revenue"]])
    ws_sum.append(["목표 달성률(%)", m["achievement_rate"]])
    ws_sum.append(["전기 대비(%)",   m["growth_rate"]])
    ws_sum.append(["진행 딜 수",     m["deal_count"]])
    ws_sum.append(["수주 건수",      m["win_count"]])
    ws_sum.append(["수주율(%)",      m["win_rate"]])
    ws_sum.append([])
    ws_sum.append(["팀장 보고 요약", report.get("summary", "")])
    ws_sum.append(["달성률 코멘트",   report.get("achievement_comment", "")])
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 60

    # 2) 파이프라인 시트
    ws_pipe = wb.create_sheet("파이프라인")
    _write_header(ws_pipe, ["단계", "건수", "금액(원)"])
    for p in report["pipeline"]:
        ws_pipe.append([p["stage"], p["count"], p["amount"]])
    ws_pipe.append([])
    ws_pipe.append(["전환율 (인접 단계)"])
    for cr in report["conversion_rates"]:
        ws_pipe.append([f"{cr['from']} → {cr['to']}", f"{cr['rate']}%"])
    ws_pipe.append([])
    ws_pipe.append(["인사이트", report.get("pipeline_insight", "")])
    ws_pipe.column_dimensions["A"].width = 24
    ws_pipe.column_dimensions["B"].width = 18
    ws_pipe.column_dimensions["C"].width = 18

    # 3) 팀원 시트
    ws_mem = wb.create_sheet("팀원")
    _write_header(ws_mem, ["이름", "매출(원)", "수주", "진행"])
    for mem in report["members"]:
        ws_mem.append([mem["name"], mem["revenue"], mem["wins"], mem["deals"]])
    ws_mem.column_dimensions["A"].width = 18

    # 4) 이상감지 시트
    ws_anom = wb.create_sheet("이상감지")
    _write_header(ws_anom, ["유형", "항목", "세부", "심각도", "원인 추정"])
    for a in report.get("anomalies", []):
        ws_anom.append([
            a.get("type", ""),
            a.get("item", ""),
            a.get("detail", ""),
            a.get("severity", ""),
            a.get("cause", ""),
        ])
    for col, width in zip("ABCDE", [10, 28, 38, 10, 50]):
        ws_anom.column_dimensions[col].width = width

    # 5) AI 인사이트 시트
    ws_ai = wb.create_sheet("AI인사이트")
    _write_header(ws_ai, ["항목", "내용"])
    ws_ai.append(["최고 실적",   report.get("top_performer", "")])
    ws_ai.append(["리스크 딜",   report.get("risk_deals", "")])
    ws_ai.append([])
    ws_ai.append(["액션 추천"])
    for i, rec in enumerate(report.get("recommendations", []), start=1):
        ws_ai.append([f"{i}", rec])
    ws_ai.column_dimensions["A"].width = 12
    ws_ai.column_dimensions["B"].width = 80

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
