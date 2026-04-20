"""
재무 라우터 — /api/finance/*
"""
import json
import uuid
from datetime import date
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, HTTPException, UploadFile, File, Query, Header
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from openai import OpenAI

from services.finance.finance_ocr_service import analyze_receipt
from services.finance.finance_account_service import suggest_account_code
from database import get_connection
from config import settings

openai_client = OpenAI(api_key=settings.openai_api_key)

router = APIRouter()


def _ensure_receipt_images_table():
    """앱 시작 시 receipt_images 테이블이 없으면 생성"""
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            CREATE TABLE IF NOT EXISTS receipt_images (
                id           UUID PRIMARY KEY,
                filename     TEXT,
                content_type TEXT,
                image_data   BYTEA,
                created_at   TIMESTAMP DEFAULT NOW()
            )
        """)
    finally:
        cur.close()
        conn.close()


# 모듈 임포트 시점에 테이블 보장
try:
    _ensure_receipt_images_table()
except Exception:
    pass  # DB 미연결 환경에서도 임포트는 허용


# ──────────────────────────────────────────────────────────────
# 요청 스키마
# ──────────────────────────────────────────────────────────────
class TransactionItem(BaseModel):
    item: str
    amount: int
    tax_amount: int
    account_code: str = "기타비용"
    memo: str = ""
    confidence: float = 0.0


class SaveTransactionsRequest(BaseModel):
    receipt_date: Optional[str] = None
    vendor: str = ""
    image_path: Optional[str] = None   # OCR 저장 시 반환된 서버 경로
    department: Optional[str] = None   # 로그인 세션에서 전달
    emp_id: Optional[str] = None       # 로그인 세션에서 전달
    items: list[TransactionItem]


class SuggestAccountRequest(BaseModel):
    vendor: str = ""
    notes: str = ""


class UpdateTransactionRequest(BaseModel):
    account_code: Optional[str] = None
    amount: Optional[int] = None
    tax_amount: Optional[int] = None
    memo: Optional[str] = None


# ──────────────────────────────────────────────────────────────
# POST /api/finance/ocr  — 이미지 저장 + OCR 분석 + 중복 탐지
# ──────────────────────────────────────────────────────────────
@router.post("/ocr")
async def ocr(file: UploadFile = File(...)):
    """
    영수증 파일을 receipt_images 테이블에 저장하고 AI로 분석합니다.
    DB 저장은 하지 않습니다. 중복 의심 항목은 is_duplicate: true로 표시됩니다.
    """
    file_bytes = await file.read()
    content_type = file.content_type or "image/jpeg"

    if not file_bytes:
        raise HTTPException(status_code=400, detail="빈 파일입니다.")

    # 이미지를 receipt_images 테이블에 저장
    image_id = str(uuid.uuid4())
    filename = file.filename or "receipt"

    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO receipt_images (id, filename, content_type, image_data) VALUES (%s, %s, %s, %s)",
            (image_id, filename, content_type, file_bytes),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"이미지 저장 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    # OCR 분석
    try:
        result = analyze_receipt(file_bytes, content_type)
    except Exception as e:
        # 분석 실패 시 저장된 이미지 롤백
        conn2 = get_connection()
        cur2  = conn2.cursor()
        try:
            cur2.execute("DELETE FROM receipt_images WHERE id = %s", (image_id,))
        finally:
            cur2.close()
            conn2.close()
        raise HTTPException(status_code=502, detail=f"AI 분석 실패: {str(e)}")

    receipt_date = result.get("receipt_date") or str(date.today())
    vendor       = result.get("vendor", "")
    items_raw    = result.get("items", [])

    if not items_raw:
        # 항목 인식 실패 시 이미지 롤백
        conn3 = get_connection()
        cur3  = conn3.cursor()
        try:
            cur3.execute("DELETE FROM receipt_images WHERE id = %s", (image_id,))
        finally:
            cur3.close()
            conn3.close()
        raise HTTPException(status_code=422, detail="영수증에서 항목을 인식하지 못했습니다.")

    # 중복 탐지 — vendor + receipt_date + amount 동시 일치 여부 조회
    dup_amounts: set[int] = set()
    if vendor and receipt_date:
        amounts = [int(it.get("amount", 0)) for it in items_raw]
        if amounts:
            conn = get_connection()
            cur  = conn.cursor()
            try:
                placeholders = ", ".join(["%s"] * len(amounts))
                cur.execute(
                    f"""
                    SELECT DISTINCT amount FROM finance_transactions
                    WHERE vendor = %s AND receipt_date = %s
                      AND amount IN ({placeholders})
                    """,
                    [vendor, receipt_date] + amounts,
                )
                dup_amounts = {r[0] for r in cur.fetchall()}
            finally:
                cur.close()
                conn.close()

    items = [
        {
            "item":         it.get("item", ""),
            "amount":       int(it.get("amount", 0)),
            "tax_amount":   int(it.get("tax_amount", 0)),
            "total_amount": int(it.get("amount", 0)) + int(it.get("tax_amount", 0)),
            "account_code": it.get("account_code", "기타비용"),
            "vendor":       vendor,
            "memo":         it.get("memo", ""),
            "confidence":   float(it.get("confidence", 0)),
            "is_duplicate": int(it.get("amount", 0)) in dup_amounts,
        }
        for it in items_raw
    ]

    return {
        "receipt_date":   receipt_date,
        "vendor":         vendor,
        "image_path":     image_id,   # UUID 문자열 (기존 필드명 유지)
        "items":          items,
        "has_duplicates": len(dup_amounts) > 0,
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/transactions  — 분석 결과를 DB에 저장
# ──────────────────────────────────────────────────────────────
@router.post("/transactions", status_code=201)
def save_transactions(body: SaveTransactionsRequest):
    """
    OCR 분석 결과를 DB에 저장합니다. status 기본값은 'pending'.
    """
    if not body.items:
        raise HTTPException(status_code=400, detail="저장할 항목이 없습니다.")

    receipt_date = body.receipt_date or str(date.today())
    vendor       = body.vendor
    saved        = []

    conn = get_connection()
    cur  = conn.cursor()

    try:
        for item in body.items:
            cur.execute(
                """
                INSERT INTO finance_transactions
                    (receipt_date, item, amount, tax_amount,
                     account_code, vendor, memo, ai_confidence, raw_json, image_path,
                     department, emp_id)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, total_amount, created_at
                """,
                (
                    receipt_date,
                    item.item,
                    item.amount,
                    item.tax_amount,
                    item.account_code,
                    vendor,
                    item.memo,
                    item.confidence,
                    json.dumps(item.model_dump(), ensure_ascii=False),
                    body.image_path,
                    body.department,
                    body.emp_id,
                ),
            )
            row = cur.fetchone()
            saved.append({
                "id":           row[0],
                "receipt_date": receipt_date,
                "item":         item.item,
                "amount":       item.amount,
                "tax_amount":   item.tax_amount,
                "total_amount": row[1],
                "account_code": item.account_code,
                "vendor":       vendor,
                "memo":         item.memo,
                "confidence":   item.confidence,
                "image_path":   body.image_path,
                "department":   body.department,
                "emp_id":       body.emp_id,
                "status":       "pending",
                "created_at":   str(row[2]),
            })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB 저장 실패: {str(e)}")

    finally:
        cur.close()
        conn.close()

    return {"saved": saved}


# ──────────────────────────────────────────────────────────────
# GET /api/finance/receipts/{image_id}  — DB에서 영수증 이미지 반환
# ──────────────────────────────────────────────────────────────
@router.get("/receipts/{image_id}")
def get_receipt_image(image_id: str):
    """
    receipt_images 테이블에서 이미지를 조회해 바이너리로 반환합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "SELECT image_data, content_type FROM receipt_images WHERE id = %s",
            (image_id,),
        )
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="이미지를 찾을 수 없습니다.")

    image_data, content_type = row
    return Response(
        content=bytes(image_data),
        media_type=content_type or "image/jpeg",
    )


# ──────────────────────────────────────────────────────────────
# DELETE /api/finance/transactions/{id}  — 전표 삭제
# ──────────────────────────────────────────────────────────────
@router.delete("/transactions/{transaction_id}")
def delete_transaction(transaction_id: int):
    """
    전표를 삭제합니다. 해당 image_id를 참조하는 전표가 더 없으면
    receipt_images도 함께 삭제합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        # 삭제 전 image_path(UUID) 조회
        cur.execute(
            "SELECT image_path FROM finance_transactions WHERE id = %s",
            (transaction_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="전표를 찾을 수 없습니다.")

        image_id = row[0]

        # 전표 삭제
        cur.execute("DELETE FROM finance_transactions WHERE id = %s", (transaction_id,))

        # 동일 image_id를 참조하는 다른 전표가 없으면 이미지도 삭제
        if image_id:
            cur.execute(
                "SELECT COUNT(*) FROM finance_transactions WHERE image_path = %s",
                (image_id,),
            )
            ref_count = cur.fetchone()[0]
            if ref_count == 0:
                cur.execute("DELETE FROM receipt_images WHERE id = %s", (image_id,))

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"삭제 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {"deleted": True}


# ──────────────────────────────────────────────────────────────
# GET /api/finance/transactions/export  — 확정 전표 엑셀 다운로드
# (주의: /{id} 라우트보다 먼저 선언해야 라우팅 충돌 없음)
# ──────────────────────────────────────────────────────────────
@router.get("/transactions/export")
def export_confirmed_excel():
    """
    status = 'confirmed' 인 전표만 엑셀(.xlsx)로 다운로드합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            """
            SELECT id, receipt_date, item, amount, tax_amount, total_amount,
                   account_code, vendor, memo, ai_confidence, created_at
            FROM finance_transactions
            WHERE status = 'confirmed'
            ORDER BY receipt_date DESC, id DESC
            """,
        )
        rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    # 엑셀 파일 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "확정 전표"

    # 헤더 스타일
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF")
    headers = ["ID", "날짜", "항목", "공급가액", "부가세", "합계", "계정과목", "거래처", "적요", "신뢰도", "등록일"]
    col_widths = [6, 12, 30, 12, 10, 12, 14, 20, 25, 8, 18]

    for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    # 데이터 행
    for row in rows:
        ws.append([
            row[0],          # id
            str(row[1]),     # receipt_date
            row[2],          # item
            row[3],          # amount
            row[4],          # tax_amount
            row[5],          # total_amount
            row[6],          # account_code
            row[7] or "",    # vendor
            row[8] or "",    # memo
            float(row[9]) if row[9] else None,  # ai_confidence
            str(row[10]),    # created_at
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename*=UTF-8''confirmed_transactions.xlsx"},
    )


# ──────────────────────────────────────────────────────────────
# GET /api/finance/transactions  — DB 전표 내역 조회
# ──────────────────────────────────────────────────────────────
@router.get("/transactions")
def list_transactions(
    limit:        int            = Query(default=50, le=200),
    offset:       int            = Query(default=0, ge=0),
    account_code: Optional[str]  = Query(default=None),
    status:       Optional[str]  = Query(default=None),
    date_from:    Optional[str]  = Query(default=None),
    date_to:      Optional[str]  = Query(default=None),
    department:   Optional[str]  = Query(default=None),
):
    """
    DB에 저장된 전표 목록을 반환합니다.

    Query params:
      limit        (int, default 50)
      offset       (int, default 0)
      account_code (str, optional)
      status       (str, optional) — 'pending' | 'confirmed'
      date_from    (str YYYY-MM-DD, optional)
      date_to      (str YYYY-MM-DD, optional)
      department   (str, optional) — 부서별 격리 조회
    """
    where_clauses = []
    params = []

    if department:
        where_clauses.append("department = %s")
        params.append(department)
    if account_code:
        where_clauses.append("account_code = %s")
        params.append(account_code)
    if status:
        where_clauses.append("status = %s")
        params.append(status)
    if date_from:
        where_clauses.append("receipt_date >= %s")
        params.append(date_from)
    if date_to:
        where_clauses.append("receipt_date <= %s")
        params.append(date_to)

    where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(f"SELECT COUNT(*) FROM finance_transactions {where_sql}", params)
        total = cur.fetchone()[0]

        cur.execute(
            f"""
            SELECT id, receipt_date, item, amount, tax_amount, total_amount,
                   account_code, vendor, memo, ai_confidence, status, image_path, created_at,
                   department, emp_id
            FROM finance_transactions
            {where_sql}
            ORDER BY created_at DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"조회 실패: {str(e)}")

    finally:
        cur.close()
        conn.close()

    return {
        "total": total,
        "items": [
            {
                "id":           r[0],
                "receipt_date": str(r[1]),
                "item":         r[2],
                "amount":       r[3],
                "tax_amount":   r[4],
                "total_amount": r[5],
                "account_code": r[6],
                "vendor":       r[7],
                "memo":         r[8],
                "confidence":   float(r[9]) if r[9] else None,
                "status":       r[10],
                "image_path":   r[11],
                "created_at":   str(r[12]),
                "department":   r[13],
                "emp_id":       r[14],
            }
            for r in rows
        ],
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/suggest-account  — AI 계정과목 추천
# ──────────────────────────────────────────────────────────────
@router.post("/suggest-account")
def suggest_account(body: SuggestAccountRequest):
    """
    가맹점명과 지출내역을 바탕으로 AI가 계정과목을 추천합니다.
    """
    try:
        code = suggest_account_code(body.vendor, body.notes)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 추천 실패: {str(e)}")
    return {"account_code": code}


# ──────────────────────────────────────────────────────────────
# PUT /api/finance/transactions/{id}  — 전표 수정
# ──────────────────────────────────────────────────────────────
@router.put("/transactions/{transaction_id}")
def update_transaction(transaction_id: int, body: UpdateTransactionRequest):
    """
    계정과목, 금액, 부가세, 적요를 수정합니다.
    """
    set_clauses = []
    params      = []

    if body.account_code is not None:
        set_clauses.append("account_code = %s"); params.append(body.account_code)
    if body.amount is not None:
        set_clauses.append("amount = %s");       params.append(body.amount)
    if body.tax_amount is not None:
        set_clauses.append("tax_amount = %s");   params.append(body.tax_amount)
    if body.memo is not None:
        set_clauses.append("memo = %s");         params.append(body.memo)

    if not set_clauses:
        raise HTTPException(status_code=400, detail="수정할 항목이 없습니다.")

    params.append(transaction_id)
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            f"""
            UPDATE finance_transactions
               SET {', '.join(set_clauses)}, updated_at = NOW()
             WHERE id = %s
            RETURNING id, account_code, amount, tax_amount, total_amount, memo, updated_at
            """,
            params,
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="전표를 찾을 수 없습니다.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"수정 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {
        "id":           row[0],
        "account_code": row[1],
        "amount":       row[2],
        "tax_amount":   row[3],
        "total_amount": row[4],
        "memo":         row[5],
        "updated_at":   str(row[6]),
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/transactions/{id}/confirm  — 최종 확정
# ──────────────────────────────────────────────────────────────
@router.post("/transactions/{transaction_id}/confirm")
def confirm_transaction(transaction_id: int):
    """
    전표 status를 'confirmed'로 업데이트합니다.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            "UPDATE finance_transactions SET status = 'confirmed', updated_at = NOW() WHERE id = %s RETURNING id, status",
            (transaction_id,),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="전표를 찾을 수 없습니다.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"확정 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {"id": row[0], "status": row[1]}


# ──────────────────────────────────────────────────────────────
# 재무팀 권한 검증 헬퍼 — employee_id로 DB에서 부서 확인
# ──────────────────────────────────────────────────────────────
def _verify_treasury_access(employee_id: Optional[str]):
    """재무팀 또는 admin(position='대표이사') 여부를 DB에서 직접 확인"""
    if not employee_id:
        raise HTTPException(status_code=401, detail="인증 정보가 없습니다.")

    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            "SELECT department, position FROM info_employees WHERE employee_id = %s AND is_verified = TRUE",
            (employee_id.strip(),),
        )
        row = cur.fetchone()
    finally:
        cur.close()
        conn.close()

    if not row:
        raise HTTPException(status_code=401, detail="유효하지 않은 사원입니다.")

    department, position = row[0], row[1]
    # 재무/회계팀, 재무팀 또는 대표이사(admin)만 접근 허용
    if department not in ("재무팀", "재무/회계팀") and position != "대표이사":
        raise HTTPException(status_code=403, detail="재무/회계팀 전용 페이지입니다.")


# ──────────────────────────────────────────────────────────────
# GET /api/finance/stats  — 재무 대시보드 통계 (재무팀 전용)
# ──────────────────────────────────────────────────────────────
@router.get("/stats")
def get_finance_stats(
    year:        int            = Query(default=2026),
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    부서별 지출 집계, 월별 지출 추이, 당월/전월 비교, 감사 위험 로그를 반환합니다.
    재무팀 or 대표이사만 접근 가능합니다.
    """
    _verify_treasury_access(x_employee_id)

    conn = get_connection()
    cur  = conn.cursor()

    try:
        # ① 부서별 지출 합계 + 예산 조인 (해당 연도)
        cur.execute(
            """
            SELECT
                t.department,
                COALESCE(SUM(t.total_amount), 0)          AS total_spent,
                COALESCE(b.budget_total, 0)                AS budget_amount
            FROM finance_transactions t
            LEFT JOIN (
                SELECT department, SUM(budget_amount) AS budget_total
                FROM finance_budgets
                WHERE fiscal_year = %s
                GROUP BY department
            ) b ON t.department = b.department
            WHERE EXTRACT(YEAR FROM t.receipt_date) = %s
            GROUP BY t.department, b.budget_total
            ORDER BY total_spent DESC
            """,
            (year, year),
        )
        dept_rows = cur.fetchall()

        dept_stats = [
            {
                "department":    r[0] or "미분류",
                "total_spent":   int(r[1]),
                "budget_amount": int(r[2]),
                # 예산 대비 집행률 (예산 0이면 None)
                "execution_rate": round(int(r[1]) / int(r[2]) * 100, 1) if int(r[2]) > 0 else None,
            }
            for r in dept_rows
        ]

        # ② 월별 지출 추이 (해당 연도)
        cur.execute(
            """
            SELECT
                EXTRACT(MONTH FROM receipt_date)::int AS month,
                COALESCE(SUM(total_amount), 0)        AS monthly_spent
            FROM finance_transactions
            WHERE EXTRACT(YEAR FROM receipt_date) = %s
            GROUP BY month
            ORDER BY month
            """,
            (year,),
        )
        monthly_rows = cur.fetchall()
        monthly_data = {r[0]: int(r[1]) for r in monthly_rows}
        # 1~12월 전체 채우기 (데이터 없는 달 = 0)
        monthly_stats = [
            {"month": f"{m}월", "spent": monthly_data.get(m, 0)}
            for m in range(1, 13)
        ]

        # ③ 당월 총 지출 & 전월 총 지출
        cur.execute(
            """
            SELECT
                COALESCE(SUM(CASE WHEN DATE_TRUNC('month', receipt_date) = DATE_TRUNC('month', CURRENT_DATE)
                                  THEN total_amount ELSE 0 END), 0) AS this_month,
                COALESCE(SUM(CASE WHEN DATE_TRUNC('month', receipt_date) = DATE_TRUNC('month', CURRENT_DATE) - INTERVAL '1 month'
                                  THEN total_amount ELSE 0 END), 0) AS last_month
            FROM finance_transactions
            """,
        )
        month_row  = cur.fetchone()
        this_month = int(month_row[0])
        last_month = int(month_row[1])
        # 전월 대비 증감률 계산
        mom_change = None
        if last_month > 0:
            mom_change = round((this_month - last_month) / last_month * 100, 1)

        # ④ 감사 위험 로그 — risk_level = 'danger' 미확인 건수
        cur.execute(
            """
            SELECT COUNT(*) FROM finance_audit_logs
            WHERE risk_level = 'danger' AND is_confirmed = FALSE
            """
        )
        danger_count = int(cur.fetchone()[0])

        # ⑤ 연간 총 예산 합계
        cur.execute(
            "SELECT COALESCE(SUM(budget_amount), 0) FROM finance_budgets WHERE fiscal_year = %s",
            (year,),
        )
        total_budget = int(cur.fetchone()[0])

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"통계 조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {
        "year":          year,
        "total_budget":  total_budget,
        "this_month":    this_month,
        "last_month":    last_month,
        "mom_change":    mom_change,   # 전월 대비 증감률(%)
        "danger_count":  danger_count,
        "dept_stats":    dept_stats,
        "monthly_stats": monthly_stats,
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/report  — CFO AI 리포트 (gpt-4o-mini)
# ──────────────────────────────────────────────────────────────
class ReportRequest(BaseModel):
    year: int = 2026


@router.post("/report")
def generate_cfo_report(
    body: ReportRequest,
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    부서별 통계 데이터를 OpenAI에 전달하여 CFO 관점의 예산 분석 리포트를 생성합니다.
    재무팀 or 대표이사만 접근 가능합니다.
    """
    _verify_treasury_access(x_employee_id)

    # 통계 데이터 재조회 (stats 엔드포인트 로직 재사용)
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
                t.department,
                COALESCE(SUM(t.total_amount), 0)          AS total_spent,
                COALESCE(b.budget_total, 0)                AS budget_amount
            FROM finance_transactions t
            LEFT JOIN (
                SELECT department, SUM(budget_amount) AS budget_total
                FROM finance_budgets
                WHERE fiscal_year = %s
                GROUP BY department
            ) b ON t.department = b.department
            WHERE EXTRACT(YEAR FROM t.receipt_date) = %s
            GROUP BY t.department, b.budget_total
            ORDER BY total_spent DESC
            """,
            (body.year, body.year),
        )
        dept_rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"데이터 조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    dept_data = [
        {
            "부서":          r[0] or "미분류",
            "집행금액(원)":   int(r[1]),
            "예산(원)":      int(r[2]),
            "집행률(%)":     round(int(r[1]) / int(r[2]) * 100, 1) if int(r[2]) > 0 else "예산 없음",
        }
        for r in dept_rows
    ]

    prompt = f"""
다음은 {body.year}년 기업 부서별 예산 집행 현황입니다:

{json.dumps(dept_data, ensure_ascii=False, indent=2)}

당신은 대기업 CFO입니다. 위 데이터를 분석하여 다음을 한국어로 작성하세요:
1. 예산 초과 위험이 있는 부서 지목 및 이유
2. 지출 패턴에서 발견된 이상 징후
3. 각 부서에 대한 구체적인 예산 관리 조언
4. 전체적인 재무 건전성 평가

간결하고 명확하게, 실무 CFO 어투로 작성하세요. 총 400자 이내.
"""

    try:
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 대기업 CFO로서 재무 데이터를 분석하고 경영진에게 보고하는 전문가입니다."},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=600,
            temperature=0.4,
        )
        report_text = response.choices[0].message.content.strip()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 리포트 생성 실패: {str(e)}")

    return {"report": report_text, "year": body.year}


# ──────────────────────────────────────────────────────────────
# POST /api/finance/seed  — 테스트용 목 데이터 삽입 (개발 전용)
# ──────────────────────────────────────────────────────────────
@router.post("/seed", status_code=201)
def seed_finance_data():
    """
    각 부서별 샘플 지출 내역과 예산 데이터를 DB에 삽입합니다.
    이미 데이터가 있으면 중복 삽입하지 않습니다.
    """
    from datetime import timedelta
    import random

    # ── 시드 트랜잭션 데이터 ──────────────────────────────────
    SEED_TRANSACTIONS = [
        # 인사팀
        {"department": "인사팀", "item": "임직원 식대 지원",      "amount": 280000,  "tax_amount": 28000,  "account_code": "복리후생비",  "vendor": "구내식당",      "receipt_date": "2026-01-15", "ai_confidence": 96.5},
        {"department": "인사팀", "item": "채용 공고 게재비",       "amount": 550000,  "tax_amount": 55000,  "account_code": "광고선전비",  "vendor": "잡코리아",      "receipt_date": "2026-01-22", "ai_confidence": 91.2},
        {"department": "인사팀", "item": "교육 훈련 세미나",       "amount": 1200000, "tax_amount": 120000, "account_code": "교육훈련비",  "vendor": "HRD코리아",     "receipt_date": "2026-02-05", "ai_confidence": 88.7},
        {"department": "인사팀", "item": "사무용품 구매",          "amount": 95000,   "tax_amount": 9500,   "account_code": "소모품비",   "vendor": "오피스디포",    "receipt_date": "2026-02-18", "ai_confidence": 99.1},
        {"department": "인사팀", "item": "임직원 건강검진",        "amount": 3200000, "tax_amount": 0,      "account_code": "복리후생비",  "vendor": "강남세브란스",  "receipt_date": "2026-03-10", "ai_confidence": 94.3},
        {"department": "인사팀", "item": "노무 컨설팅 수수료",     "amount": 800000,  "tax_amount": 80000,  "account_code": "수수료비용",  "vendor": "노무법인한울",  "receipt_date": "2026-03-25", "ai_confidence": 87.6},

        # 마케팅팀
        {"department": "마케팅팀", "item": "SNS 광고 집행",        "amount": 5000000, "tax_amount": 500000, "account_code": "광고선전비",  "vendor": "메타코리아",    "receipt_date": "2026-01-08", "ai_confidence": 95.0},
        {"department": "마케팅팀", "item": "홍보물 인쇄",          "amount": 320000,  "tax_amount": 32000,  "account_code": "도서인쇄비",  "vendor": "프린팅아이",    "receipt_date": "2026-01-20", "ai_confidence": 92.4},
        {"department": "마케팅팀", "item": "브랜드 콘텐츠 제작",   "amount": 2800000, "tax_amount": 280000, "account_code": "광고선전비",  "vendor": "크리에이티브랩", "receipt_date": "2026-02-14", "ai_confidence": 89.8},
        {"department": "마케팅팀", "item": "전시회 참가비",        "amount": 1500000, "tax_amount": 150000, "account_code": "광고선전비",  "vendor": "코엑스",        "receipt_date": "2026-02-28", "ai_confidence": 93.1},
        {"department": "마케팅팀", "item": "고객 설문 툴 구독료",  "amount": 450000,  "tax_amount": 45000,  "account_code": "수수료비용",  "vendor": "서베이몽키",    "receipt_date": "2026-03-05", "ai_confidence": 97.2},
        {"department": "마케팅팀", "item": "인플루언서 마케팅",    "amount": 3500000, "tax_amount": 350000, "account_code": "광고선전비",  "vendor": "인플루언서팀",  "receipt_date": "2026-03-20", "ai_confidence": 85.5},

        # 개발팀
        {"department": "개발팀", "item": "AWS 클라우드 이용료",    "amount": 4200000, "tax_amount": 420000, "account_code": "임차료",     "vendor": "Amazon AWS",    "receipt_date": "2026-01-31", "ai_confidence": 98.9},
        {"department": "개발팀", "item": "GitHub Enterprise 구독", "amount": 890000,  "tax_amount": 89000,  "account_code": "수수료비용",  "vendor": "GitHub Inc.",   "receipt_date": "2026-01-15", "ai_confidence": 99.0},
        {"department": "개발팀", "item": "개발 장비 구매(모니터)", "amount": 2100000, "tax_amount": 210000, "account_code": "소모품비",   "vendor": "삼성전자",      "receipt_date": "2026-02-03", "ai_confidence": 90.5},
        {"department": "개발팀", "item": "기술 세미나 참가",       "amount": 600000,  "tax_amount": 60000,  "account_code": "교육훈련비",  "vendor": "AWS Summit",    "receipt_date": "2026-02-20", "ai_confidence": 88.3},
        {"department": "개발팀", "item": "소프트웨어 라이선스",    "amount": 3800000, "tax_amount": 380000, "account_code": "수수료비용",  "vendor": "JetBrains",     "receipt_date": "2026-03-01", "ai_confidence": 96.7},
        {"department": "개발팀", "item": "보안 취약점 점검 용역",  "amount": 5500000, "tax_amount": 550000, "account_code": "수수료비용",  "vendor": "시큐어웍스",    "receipt_date": "2026-03-18", "ai_confidence": 82.1},

        # 영업팀
        {"department": "영업팀", "item": "고객사 접대 식사",       "amount": 480000,  "tax_amount": 48000,  "account_code": "접대비",     "vendor": "강남 한정식",   "receipt_date": "2026-01-12", "ai_confidence": 94.6},
        {"department": "영업팀", "item": "출장 교통비(KTX)",       "amount": 125000,  "tax_amount": 0,      "account_code": "여비교통비",  "vendor": "한국철도공사",  "receipt_date": "2026-01-19", "ai_confidence": 99.5},
        {"department": "영업팀", "item": "고객 선물 구매",         "amount": 750000,  "tax_amount": 75000,  "account_code": "접대비",     "vendor": "신세계백화점",  "receipt_date": "2026-02-09", "ai_confidence": 91.8},
        {"department": "영업팀", "item": "영업 활동 차량 주유비",  "amount": 220000,  "tax_amount": 0,      "account_code": "여비교통비",  "vendor": "SK에너지",      "receipt_date": "2026-02-22", "ai_confidence": 97.3},
        {"department": "영업팀", "item": "파트너사 골프 접대",      "amount": 980000,  "tax_amount": 98000,  "account_code": "접대비",     "vendor": "레이크힐스CC", "receipt_date": "2026-03-08", "ai_confidence": 86.4},
        {"department": "영업팀", "item": "영업 자료 인쇄",         "amount": 85000,   "tax_amount": 8500,   "account_code": "도서인쇄비",  "vendor": "킨코스",        "receipt_date": "2026-03-22", "ai_confidence": 98.2},

        # 운영팀
        {"department": "운영팀", "item": "사무실 소모품",          "amount": 180000,  "tax_amount": 18000,  "account_code": "소모품비",   "vendor": "이마트",        "receipt_date": "2026-01-06", "ai_confidence": 95.8},
        {"department": "운영팀", "item": "건물 시설 보수 공사",    "amount": 3200000, "tax_amount": 320000, "account_code": "수수료비용",  "vendor": "현대건설",      "receipt_date": "2026-01-25", "ai_confidence": 83.7},
        {"department": "운영팀", "item": "복합기 임대료",          "amount": 250000,  "tax_amount": 25000,  "account_code": "임차료",     "vendor": "신도리코",      "receipt_date": "2026-02-01", "ai_confidence": 99.3},
        {"department": "운영팀", "item": "인터넷/전화 통신비",     "amount": 320000,  "tax_amount": 32000,  "account_code": "통신비",     "vendor": "KT",            "receipt_date": "2026-02-28", "ai_confidence": 98.7},
        {"department": "운영팀", "item": "청소 용역비",            "amount": 550000,  "tax_amount": 55000,  "account_code": "수수료비용",  "vendor": "깨끗한나라",    "receipt_date": "2026-03-07", "ai_confidence": 92.0},
        {"department": "운영팀", "item": "직원 간식 구매",         "amount": 150000,  "tax_amount": 15000,  "account_code": "복리후생비",  "vendor": "GS25",          "receipt_date": "2026-03-28", "ai_confidence": 96.1},
    ]

    # ── 시드 예산 데이터 ──────────────────────────────────────
    SEED_BUDGETS = [
        # 인사팀
        {"fiscal_year": 2026, "department": "인사팀",   "account_code": "복리후생비",  "budget_amount": 8000000},
        {"fiscal_year": 2026, "department": "인사팀",   "account_code": "교육훈련비",  "budget_amount": 5000000},
        {"fiscal_year": 2026, "department": "인사팀",   "account_code": "광고선전비",  "budget_amount": 3000000},
        {"fiscal_year": 2026, "department": "인사팀",   "account_code": "소모품비",   "budget_amount": 1000000},
        {"fiscal_year": 2026, "department": "인사팀",   "account_code": "수수료비용",  "budget_amount": 2000000},
        # 마케팅팀
        {"fiscal_year": 2026, "department": "마케팅팀", "account_code": "광고선전비",  "budget_amount": 20000000},
        {"fiscal_year": 2026, "department": "마케팅팀", "account_code": "수수료비용",  "budget_amount": 3000000},
        {"fiscal_year": 2026, "department": "마케팅팀", "account_code": "도서인쇄비",  "budget_amount": 1000000},
        # 개발팀
        {"fiscal_year": 2026, "department": "개발팀",   "account_code": "임차료",     "budget_amount": 15000000},
        {"fiscal_year": 2026, "department": "개발팀",   "account_code": "수수료비용",  "budget_amount": 25000000},
        {"fiscal_year": 2026, "department": "개발팀",   "account_code": "소모품비",   "budget_amount": 5000000},
        {"fiscal_year": 2026, "department": "개발팀",   "account_code": "교육훈련비",  "budget_amount": 3000000},
        # 영업팀
        {"fiscal_year": 2026, "department": "영업팀",   "account_code": "접대비",     "budget_amount": 5000000},
        {"fiscal_year": 2026, "department": "영업팀",   "account_code": "여비교통비",  "budget_amount": 3000000},
        {"fiscal_year": 2026, "department": "영업팀",   "account_code": "도서인쇄비",  "budget_amount": 500000},
        # 운영팀
        {"fiscal_year": 2026, "department": "운영팀",   "account_code": "소모품비",   "budget_amount": 2000000},
        {"fiscal_year": 2026, "department": "운영팀",   "account_code": "수수료비용",  "budget_amount": 6000000},
        {"fiscal_year": 2026, "department": "운영팀",   "account_code": "임차료",     "budget_amount": 3000000},
        {"fiscal_year": 2026, "department": "운영팀",   "account_code": "통신비",     "budget_amount": 1500000},
        {"fiscal_year": 2026, "department": "운영팀",   "account_code": "복리후생비",  "budget_amount": 2000000},
    ]

    conn = get_connection()
    cur  = conn.cursor()

    try:
        # 트랜잭션 데이터 삽입 (중복 방지: vendor+receipt_date+item 동시 일치 시 스킵)
        inserted_tx = 0
        for tx in SEED_TRANSACTIONS:
            cur.execute(
                """
                SELECT COUNT(*) FROM finance_transactions
                WHERE vendor = %s AND receipt_date = %s AND item = %s
                """,
                (tx["vendor"], tx["receipt_date"], tx["item"]),
            )
            if cur.fetchone()[0] == 0:
                cur.execute(
                    """
                    INSERT INTO finance_transactions
                        (receipt_date, item, amount, tax_amount, account_code,
                         department, vendor, ai_confidence)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        tx["receipt_date"], tx["item"], tx["amount"], tx["tax_amount"],
                        tx["account_code"], tx["department"], tx["vendor"], tx["ai_confidence"],
                    ),
                )
                inserted_tx += 1

        # 예산 데이터 삽입 (UNIQUE 제약 조건으로 중복 방지)
        inserted_bgt = 0
        for b in SEED_BUDGETS:
            cur.execute(
                """
                INSERT INTO finance_budgets (fiscal_year, department, account_code, budget_amount)
                VALUES (%s, %s, %s, %s)
                ON CONFLICT (fiscal_year, department, account_code) DO NOTHING
                """,
                (b["fiscal_year"], b["department"], b["account_code"], b["budget_amount"]),
            )
            inserted_bgt += 1

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"시드 데이터 삽입 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {
        "message":       "시드 데이터 삽입 완료",
        "transactions":  inserted_tx,
        "budgets":       inserted_bgt,
    }


# ══════════════════════════════════════════════════════════════
# 내부감사 (FDS) — /api/finance/audit/*
# ══════════════════════════════════════════════════════════════

class ConfirmAuditRequest(BaseModel):
    confirmed_by: str   # 확인자 이름 or 사번


# ──────────────────────────────────────────────────────────────
# POST /api/finance/audit/run  — 이상 지출 탐지 엔진 실행
# ──────────────────────────────────────────────────────────────
@router.post("/audit/run")
def run_audit_detection(
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    Rule-based 필터링 → AI 심층 분석 → finance_audit_logs 저장.

    탐지 규칙
    ① 동일 부서·계정과목의 이전 누적 평균 대비 150% 초과 지출
    ② 주말(토·일) 발생 고액 지출 (total_amount ≥ 500,000원)

    이미 감사 로그가 있는 전표는 중복 분석하지 않습니다.
    """
    _verify_treasury_access(x_employee_id)

    conn = get_connection()
    cur  = conn.cursor()

    try:
        # ── Rule ①: 누적 평균 대비 150% 초과 건 ─────────────
        cur.execute(
            """
            WITH tx_with_prev_avg AS (
                SELECT
                    id, receipt_date, item, amount, tax_amount, total_amount,
                    account_code, department, vendor, memo, ai_confidence,
                    AVG(total_amount) OVER (
                        PARTITION BY department, account_code
                        ORDER BY receipt_date, id
                        ROWS BETWEEN UNBOUNDED PRECEDING AND 1 PRECEDING
                    ) AS prev_avg
                FROM finance_transactions
                WHERE department IS NOT NULL AND department != ''
            )
            SELECT
                id, receipt_date, item, amount, tax_amount, total_amount,
                account_code, department, vendor, memo, ai_confidence,
                ROUND(prev_avg) AS prev_avg,
                ROUND(total_amount::numeric / NULLIF(prev_avg, 0) * 100, 1) AS pct_of_prev_avg,
                'amount_anomaly' AS rule_type
            FROM tx_with_prev_avg
            WHERE prev_avg IS NOT NULL
              AND prev_avg > 0
              AND total_amount > prev_avg * 1.5
              AND NOT EXISTS (
                SELECT 1 FROM finance_audit_logs al WHERE al.transaction_id = tx_with_prev_avg.id
              )
            ORDER BY pct_of_prev_avg DESC
            """
        )
        rule1_rows = cur.fetchall()

        # ── Rule ②: 주말 고액 지출 건 ────────────────────────
        cur.execute(
            """
            SELECT
                id, receipt_date, item, amount, tax_amount, total_amount,
                account_code, department, vendor, memo, ai_confidence,
                NULL AS prev_avg,
                NULL AS pct_of_prev_avg,
                'weekend_high_spend' AS rule_type
            FROM finance_transactions
            WHERE EXTRACT(DOW FROM receipt_date) IN (0, 6)   -- 0=일요일, 6=토요일
              AND total_amount >= 500000
              AND department IS NOT NULL AND department != ''
              AND NOT EXISTS (
                SELECT 1 FROM finance_audit_logs al WHERE al.transaction_id = finance_transactions.id
              )
            ORDER BY total_amount DESC
            """
        )
        rule2_rows = cur.fetchall()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"탐지 쿼리 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    # 중복 제거: rule1 + rule2 합산 후 id 기준 유니크
    seen_ids: set[int] = set()
    suspicious: list[dict] = []

    def _row_to_dict(r, rule_type: str) -> dict:
        return {
            "id":            r[0],
            "receipt_date":  str(r[1]),
            "item":          r[2],
            "amount":        int(r[3]),
            "tax_amount":    int(r[4]),
            "total_amount":  int(r[5]),
            "account_code":  r[6],
            "department":    r[7],
            "vendor":        r[8] or "",
            "memo":          r[9] or "",
            "ai_confidence": float(r[10]) if r[10] else None,
            "prev_avg":      int(r[11]) if r[11] else None,
            "pct_of_prev_avg": float(r[12]) if r[12] else None,
            "rule_type":     rule_type,
        }

    for r in rule1_rows:
        if r[0] not in seen_ids:
            seen_ids.add(r[0])
            suspicious.append(_row_to_dict(r, "amount_anomaly"))

    for r in rule2_rows:
        if r[0] not in seen_ids:
            seen_ids.add(r[0])
            suspicious.append(_row_to_dict(r, "weekend_high_spend"))

    if not suspicious:
        return {"message": "탐지된 이상 지출이 없습니다.", "analyzed": 0, "saved": 0}

    # ── AI 심층 분석 (gpt-4o-mini) ───────────────────────────
    RULE_LABEL = {
        "amount_anomaly":    "평균 대비 고액 지출",
        "weekend_high_spend": "주말 고액 지출",
    }
    AI_SYSTEM = (
        "당신은 까다로운 대기업 내부 감사역입니다. "
        "제공된 지출 내역이 회사 규정(접대비 한도 50만원, 복리후생비 연간 한도, 수의계약 기준 등)에 "
        "어긋나는지 엄격히 분석하세요. "
        "각 항목에 대해 반드시 다음 JSON 배열로만 응답하세요:\n"
        '[{"transaction_id": <int>, "risk_level": "safe"|"warning"|"danger", '
        '"violated_rule": "<위반 규정 조항 또는 null>", '
        '"ai_reason": "<판단 사유 (50자 이내)>"}]'
    )

    items_payload = []
    for s in suspicious:
        rule_label = RULE_LABEL.get(s["rule_type"], s["rule_type"])
        line = (
            f'ID:{s["id"]} | {s["department"]} | {s["account_code"]} | '
            f'{s["item"]} | {s["total_amount"]:,}원'
        )
        if s["prev_avg"]:
            line += f' | 평균대비 {s["pct_of_prev_avg"]}% (평균 {s["prev_avg"]:,}원)'
        if s["rule_type"] == "weekend_high_spend":
            line += f' | 주말 발생 ({s["receipt_date"]})'
        line += f' | 탐지규칙: {rule_label}'
        items_payload.append(line)

    user_prompt = "다음 의심 지출 내역을 분석하세요:\n\n" + "\n".join(items_payload)

    try:
        ai_resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": AI_SYSTEM},
                {"role": "user",   "content": user_prompt},
            ],
            max_tokens=1500,
            temperature=0.2,   # 일관된 판단을 위해 낮은 temperature
        )
        raw_text = ai_resp.choices[0].message.content.strip()
        # JSON 파싱 — 코드블록 제거
        raw_text = raw_text.replace("```json", "").replace("```", "").strip()
        ai_results: list[dict] = json.loads(raw_text)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 분석 실패: {str(e)}")

    # AI 결과를 transaction_id 기준 딕셔너리로 인덱싱
    ai_map: dict[int, dict] = {r["transaction_id"]: r for r in ai_results}

    # ── finance_audit_logs 저장 ───────────────────────────────
    conn2 = get_connection()
    cur2  = conn2.cursor()
    saved = 0
    saved_logs = []

    try:
        for s in suspicious:
            tid   = s["id"]
            ai    = ai_map.get(tid, {})
            level = ai.get("risk_level", "warning")
            if level not in ("safe", "warning", "danger"):
                level = "warning"

            reason  = ai.get("ai_reason", "규정 위반 여부 분석 필요")
            v_rule  = ai.get("violated_rule") or None
            raw_obj = {
                "transaction": s,
                "ai_result":   ai,
                "rule_type":   s["rule_type"],
            }

            cur2.execute(
                """
                INSERT INTO finance_audit_logs
                    (transaction_id, risk_level, violated_rule, ai_reason, raw_json)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id, transaction_id, risk_level, violated_rule, ai_reason, created_at
                """,
                (tid, level, v_rule, reason, json.dumps(raw_obj, ensure_ascii=False, default=str)),
            )
            row = cur2.fetchone()
            saved += 1
            saved_logs.append({
                "log_id":        row[0],
                "transaction_id": row[1],
                "risk_level":    row[2],
                "violated_rule": row[3],
                "ai_reason":     row[4],
                "created_at":    str(row[5]),
            })

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"감사 로그 저장 실패: {str(e)}")
    finally:
        cur2.close()
        conn2.close()

    return {
        "message":  f"이상 지출 탐지 완료",
        "analyzed": len(suspicious),
        "saved":    saved,
        "logs":     saved_logs,
    }


# ──────────────────────────────────────────────────────────────
# GET /api/finance/audit/logs  — 감사 로그 조회
# ──────────────────────────────────────────────────────────────
@router.get("/audit/logs")
def list_audit_logs(
    risk_level:   Optional[str] = Query(default=None),
    is_confirmed: Optional[bool] = Query(default=None),
    limit:        int            = Query(default=100, le=500),
    offset:       int            = Query(default=0, ge=0),
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    finance_audit_logs를 finance_transactions와 조인하여 반환합니다.
    재무팀 or 대표이사만 접근 가능합니다.
    """
    _verify_treasury_access(x_employee_id)

    where = []
    params = []

    if risk_level:
        where.append("al.risk_level = %s")
        params.append(risk_level)
    if is_confirmed is not None:
        where.append("al.is_confirmed = %s")
        params.append(is_confirmed)

    where_sql = ("WHERE " + " AND ".join(where)) if where else ""

    conn = get_connection()
    cur  = conn.cursor()

    try:
        cur.execute(
            f"SELECT COUNT(*) FROM finance_audit_logs al {where_sql}", params
        )
        total = int(cur.fetchone()[0])

        cur.execute(
            f"""
            SELECT
                al.id, al.transaction_id, al.risk_level, al.violated_rule,
                al.ai_reason, al.is_confirmed, al.confirmed_by, al.confirmed_at,
                al.created_at,
                t.receipt_date, t.item, t.amount, t.tax_amount, t.total_amount,
                t.account_code, t.department, t.vendor, t.memo
            FROM finance_audit_logs al
            LEFT JOIN finance_transactions t ON t.id = al.transaction_id
            {where_sql}
            ORDER BY
                CASE al.risk_level WHEN 'danger' THEN 1 WHEN 'warning' THEN 2 ELSE 3 END,
                al.created_at DESC
            LIMIT %s OFFSET %s
            """,
            params + [limit, offset],
        )
        rows = cur.fetchall()

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"감사 로그 조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    items = [
        {
            "id":            r[0],
            "transaction_id": r[1],
            "risk_level":    r[2],
            "violated_rule": r[3],
            "ai_reason":     r[4],
            "is_confirmed":  r[5],
            "confirmed_by":  r[6],
            "confirmed_at":  str(r[7]) if r[7] else None,
            "created_at":    str(r[8]),
            # 전표 정보
            "receipt_date":  str(r[9])  if r[9]  else None,
            "item":          r[10],
            "amount":        int(r[11]) if r[11] else None,
            "tax_amount":    int(r[12]) if r[12] else None,
            "total_amount":  int(r[13]) if r[13] else None,
            "account_code":  r[14],
            "department":    r[15],
            "vendor":        r[16],
            "memo":          r[17],
        }
        for r in rows
    ]

    return {"total": total, "items": items}


# ──────────────────────────────────────────────────────────────
# PUT /api/finance/audit/logs/{id}/confirm  — 감사 확인 처리
# ──────────────────────────────────────────────────────────────
@router.put("/audit/logs/{log_id}/confirm")
def confirm_audit_log(
    log_id: int,
    body:   ConfirmAuditRequest,
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    is_confirmed=TRUE, confirmed_by·confirmed_at 업데이트.
    재무팀 or 대표이사만 가능합니다.
    """
    _verify_treasury_access(x_employee_id)

    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute(
            """
            UPDATE finance_audit_logs
               SET is_confirmed = TRUE,
                   confirmed_by = %s,
                   confirmed_at = NOW()
             WHERE id = %s
            RETURNING id, is_confirmed, confirmed_by, confirmed_at
            """,
            (body.confirmed_by.strip(), log_id),
        )
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="감사 로그를 찾을 수 없습니다.")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"확인 처리 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    return {
        "id":           row[0],
        "is_confirmed": row[1],
        "confirmed_by": row[2],
        "confirmed_at": str(row[3]),
    }


# ──────────────────────────────────────────────────────────────
# POST /api/finance/audit/report  — 월간 감사 보고서 생성
# ──────────────────────────────────────────────────────────────
@router.post("/audit/report")
def generate_audit_report(
    x_employee_id: Optional[str] = Header(default=None, alias="x-employee-id"),
):
    """
    이번 달 Danger 감사 로그를 집계하여 GPT로 월간 감사 보고서를 생성합니다.
    재무팀 or 대표이사만 접근 가능합니다.
    """
    _verify_treasury_access(x_employee_id)

    conn = get_connection()
    cur  = conn.cursor()
    try:
        # 이번 달 전체 감사 로그 (danger + warning)
        cur.execute(
            """
            SELECT
                al.risk_level, al.violated_rule, al.ai_reason,
                al.is_confirmed, al.confirmed_by,
                t.receipt_date, t.item, t.total_amount,
                t.account_code, t.department, t.vendor
            FROM finance_audit_logs al
            LEFT JOIN finance_transactions t ON t.id = al.transaction_id
            WHERE DATE_TRUNC('month', al.created_at) = DATE_TRUNC('month', CURRENT_TIMESTAMP)
              AND al.risk_level IN ('danger', 'warning')
            ORDER BY
                CASE al.risk_level WHEN 'danger' THEN 1 ELSE 2 END,
                al.created_at DESC
            """
        )
        rows = cur.fetchall()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"데이터 조회 실패: {str(e)}")
    finally:
        cur.close()
        conn.close()

    if not rows:
        return {"report": "이번 달 탐지된 Danger/Warning 항목이 없습니다.", "danger_count": 0, "warning_count": 0}

    danger_count  = sum(1 for r in rows if r[0] == "danger")
    warning_count = sum(1 for r in rows if r[0] == "warning")

    # 보고서용 데이터 포맷팅
    items_text = []
    for i, r in enumerate(rows, 1):
        confirmed_str = f"확인 완료({r[4]})" if r[3] else "미확인"
        items_text.append(
            f"{i}. [{r[0].upper()}] {r[9]}팀 | {r[8]} | {r[7]:,}원 "
            f"({r[5]}) | 위반: {r[2]} | {confirmed_str}"
        )

    prompt = f"""
다음은 이번 달 내부 감사에서 탐지된 이상 지출 목록입니다:

{chr(10).join(items_text)}

당신은 대기업 내부감사팀장입니다. 위 데이터를 기반으로 공식 월간 감사 보고서를 한국어로 작성하세요.

보고서 형식:
1. 요약 (Danger {danger_count}건, Warning {warning_count}건 총평)
2. 주요 위험 항목 분석 (Danger 건별 핵심 요약)
3. 부서별 이상 패턴 분석
4. 권고 사항 (경영진 제출용)

간결하고 공식적인 보고서 어투로, 총 500자 이내로 작성하세요.
"""
    try:
        resp = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "당신은 대기업 내부감사팀장으로, 명확하고 간결한 공식 보고서를 작성하는 전문가입니다."},
                {"role": "user",   "content": prompt},
            ],
            max_tokens=800,
            temperature=0.3,
        )
        report_text = resp.choices[0].message.content.strip()
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"AI 보고서 생성 실패: {str(e)}")

    return {
        "report":        report_text,
        "danger_count":  danger_count,
        "warning_count": warning_count,
        "total_items":   len(rows),
    }
